"""
Ambition Campus — Base de Données Master & Enrichissement Fondations d'Entreprise
Pilier 3 : 55 Fondations d'Entreprise et Fonds de Dotation Qualifiés en France.
Génère le CSV complet et le fichier Excel formaté avec tableau de bord et filtres.
"""

import os
import re
import csv
import pandas as pd

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "prospection", "fondations")
CSV_PATH = os.path.join(OUTPUT_DIR, "fondations_database.csv")
XLSX_PATH = os.path.join(OUTPUT_DIR, "fondations_database.xlsx")
DOWNLOAD_PATH = r"C:\Users\User\Downloads\ID-NomFondation-GroupeParent-ThematiquesCibles-Nom.csv"

# ── LOT 1 : 26 PREMIÈRES FONDATIONS QUALIFIÉES (TIER 1 À TIER 4) ──
BASE_FONDATIONS_LOT1 = [
    {
        "ID": "FOND-01",
        "Nom_Fondation": "Fondation CANAL+ / Groupe Bolloré",
        "Groupe_Parent": "Groupe Canal+ / Bolloré SE",
        "Priorite": "Tier 1 - Piste Chaude",
        "Thematiques_Cibles": "Égalité des chances, Insertion jeunes, Accès culture & médias",
        "Nom_Contact": "Marine Schenfele",
        "Poste_Contact": "Déléguée Générale Fondation Canal+",
        "Email_Contact": "fondation@canal-plus.com / mecenat@bollore.com",
        "Site_Web": "https://www.canalplusgroup.com/fr/groupe/nos-engagements",
        "Lien_Depot_AAP": "Recommandation directe via équipe Canal+ (Contact chaud)",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Type_Approche": "Outreach recommandé (Email direct citant l'introducteur Canal+)",
        "Angle_Pitch_Ambition_Campus": "Partenariat 17 ans, 500+ jeunes QPV propulsés vers les filières sélectives, ROI social 1€=5,30€.",
        "Statut_Prospection": "À contacter en priorité (Email prêt)",
        "Notes_Action": "Envoyer le template d'email Bolloré avec la plaquette A4 en pièce jointe."
    },
    {
        "ID": "FOND-02",
        "Nom_Fondation": "Fondation Solidarité PwC France",
        "Groupe_Parent": "PwC France et Maghreb",
        "Priorite": "Tier 1 - Piste Chaude",
        "Thematiques_Cibles": "Égalité des chances, Éducation, Insertion économique & sociale",
        "Nom_Contact": "Adélaïde de Tourtier / Candice Galopeau",
        "Poste_Contact": "Directrice RSE & Déléguée Fondation / Resp. Engagement Sociétal",
        "Email_Contact": "adelaide.de.tourtier@pwc.com / candice.galopeau@pwc.com",
        "Site_Web": "https://www.pwc.fr/fr/fondation.html",
        "Lien_Depot_AAP": "Contact direct partenariats existants",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Type_Approche": "Upgrade de partenariat (Collaborateurs PwC déjà jurys aux oraux blancs)",
        "Angle_Pitch_Ambition_Campus": "Transformer leur implication de jury bénévole en mécénat financier officiel annuel.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Proposer une convention annuelle de mécénat avec accueil d'une masterclass dans leurs locaux."
    },
    {
        "ID": "FOND-03",
        "Nom_Fondation": "Fondation Deloitte France",
        "Groupe_Parent": "Deloitte France",
        "Priorite": "Tier 1 - Piste Chaude",
        "Thematiques_Cibles": "Éducation en zones prioritaires (ZEP/REP), Égalité des chances, Employabilité",
        "Nom_Contact": "Pôle Attractivité & RSE",
        "Poste_Contact": "Responsable Mécénat Éducation & RSE",
        "Email_Contact": "frpoleattractivite@deloitte.fr",
        "Site_Web": "https://www2.deloitte.com/fr/fr/pages/about-deloitte/articles/fondation-deloitte.html",
        "Lien_Depot_AAP": "Prix de la Fondation Deloitte & Contact direct partenariats",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Type_Approche": "Upgrade de partenariat (Collaborateurs Deloitte déjà jurys)",
        "Angle_Pitch_Ambition_Campus": "Pérennisation du pôle oraux blancs (210+ simulations) et bourses de mobilité.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Cibler le responsable mécénat éducation à la Tour Majunga / La Défense."
    },
    {
        "ID": "FOND-04",
        "Nom_Fondation": "Direction Engagement Citoyen KPMG",
        "Groupe_Parent": "KPMG France (Entreprise à Mission)",
        "Priorite": "Tier 1 - Piste Chaude",
        "Thematiques_Cibles": "Éducation, Égalité des chances (« Les Lycées de la Réussite »), Insertion",
        "Nom_Contact": "Direction de l'Engagement Citoyen",
        "Poste_Contact": "Directeur / Chargé de mission RSE & Mécénat",
        "Email_Contact": "contact via Tour Eqho (prenom.nom@kpmg.fr)",
        "Site_Web": "https://kpmg.com/fr/fr/about/engagement-citoyen.html",
        "Lien_Depot_AAP": "Programme Les Lycées de la Réussite (synergie directe)",
        "Ticket_Moyen_Estime": "10 000 € - 15 000 €",
        "Type_Approche": "Upgrade de partenariat (Collaborateurs KPMG déjà jurys)",
        "Angle_Pitch_Ambition_Campus": "Alignement total avec leur programme « Les Lycées de la Réussite » et les 36 lycées conventionnés.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Solliciter un call de 15 minutes avec le pôle Engagement Citoyen."
    },
    {
        "ID": "FOND-05",
        "Nom_Fondation": "Fondation d'Entreprise EY France",
        "Groupe_Parent": "EY France (Ernst & Young)",
        "Priorite": "Tier 1 - Piste Chaude",
        "Thematiques_Cibles": "Insertion par la formation, égalité des chances, mécénat de compétences",
        "Nom_Contact": "Fabienne Marqueste / Orane Tribouley",
        "Poste_Contact": "Déléguée Générale / Resp. accompagnement projets",
        "Email_Contact": "fabienne.marqueste@fr.ey.com / orane.tribouley@fr.ey.com",
        "Site_Web": "http://www.fondation-ey.com",
        "Lien_Depot_AAP": "Formulaire de candidature et contact direct",
        "Ticket_Moyen_Estime": "8 000 € - 15 000 €",
        "Type_Approche": "Upgrade de partenariat (Collaborateurs EY déjà jurys)",
        "Angle_Pitch_Ambition_Campus": "Co-financement des ateliers de prise de parole et préparation aux concours sélectifs.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Prendre contact avec Orane Tribouley à la Tour First (Paris La Défense)."
    },
    {
        "ID": "FOND-06",
        "Nom_Fondation": "Fondation BNP Paribas (Projet Banlieues)",
        "Groupe_Parent": "BNP Paribas",
        "Priorite": "Tier 2 - Grand Donateur",
        "Thematiques_Cibles": "Quartiers Prioritaires (QPV), Éducation, Insertion des jeunes",
        "Nom_Contact": "Isabelle Giordano",
        "Poste_Contact": "Déléguée Générale Fondation BNP Paribas",
        "Email_Contact": "via https://projet-banlieues.fondation.bnpparibas/fr/",
        "Site_Web": "https://fondation.bnpparibas",
        "Lien_Depot_AAP": "https://projet-banlieues.fondation.bnpparibas/fr/ (Campagne annuelle Jan-Mars)",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Type_Approche": "Candidature AAP officiel + Contact institutionnel",
        "Angle_Pitch_Ambition_Campus": "Ancrage exclusif en QPV/REP depuis 17 ans, +500 lycéens accompagnés, 100% bénévole.",
        "Statut_Prospection": "À préparer pour la prochaine session AAP",
        "Notes_Action": "Préparer le dossier avec les statuts, bilan comptable et attestation d'intérêt général."
    },
    {
        "ID": "FOND-07",
        "Nom_Fondation": "Fondation Société Générale (« C'est vous l'avenir »)",
        "Groupe_Parent": "Société Générale",
        "Priorite": "Tier 2 - Grand Donateur",
        "Thematiques_Cibles": "Éducation, Savoirs fondamentaux, Insertion professionnelle des jeunes <30 ans",
        "Nom_Contact": "Albane Rouvillois",
        "Poste_Contact": "Déléguée Générale & Directrice Mécénat",
        "Email_Contact": "Fondation-SocieteGenerale@socgen.com",
        "Site_Web": "https://fondation.societegenerale.com",
        "Lien_Depot_AAP": "https://submit.fondation.societegenerale.com/fr",
        "Ticket_Moyen_Estime": "15 000 € - 35 000 €",
        "Type_Approche": "Dépôt de dossier plateforme + Email d'accroche",
        "Angle_Pitch_Ambition_Campus": "Insertion par l'excellence académique : 21 admis Sciences Po, 17 Sorbonne, 13 Prépas.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Envoyer un email d'introduction à Fondation-SocieteGenerale@socgen.com."
    },
    {
        "ID": "FOND-08",
        "Nom_Fondation": "Fonds MAIF pour l'Éducation",
        "Groupe_Parent": "MAIF",
        "Priorite": "Tier 2 - Grand Donateur",
        "Thematiques_Cibles": "Éducation, Inclusion, Émancipation des publics vulnérables, Égalité des chances",
        "Nom_Contact": "Secrétariat Général du Fonds MAIF",
        "Poste_Contact": "Responsable des Appels à Projets Éducation",
        "Email_Contact": "fondsmaif@maif.fr",
        "Site_Web": "https://fonds-maif-education.fr",
        "Lien_Depot_AAP": "https://fonds-maif-education.fr/appels-a-projets",
        "Ticket_Moyen_Estime": "5 000 € - 15 000 €",
        "Type_Approche": "Candidature AAP annuel (prix départementaux & nationaux)",
        "Angle_Pitch_Ambition_Campus": "Lutte contre le déterminisme social et émancipation intellectuelle des lycéens REP.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Candidater sur le volet Île-de-France et antenne Poitiers (proximité siège Niort)."
    },
    {
        "ID": "FOND-09",
        "Nom_Fondation": "Fondation d'Entreprise Banque Populaire / BPCE",
        "Groupe_Parent": "Groupe BPCE (Banque Populaire / Caisse d'Épargne)",
        "Priorite": "Tier 2 - Grand Donateur",
        "Thematiques_Cibles": "Projets de vie, Parcours d'audace, Insertion des jeunes",
        "Nom_Contact": "Martine Tremblay / Léonard Barbu",
        "Poste_Contact": "Directrice Mécénat / Chargé de projets",
        "Email_Contact": "martine.tremblay@fnbp.banquepopulaire.fr",
        "Site_Web": "https://www.fondationbanquepopulaire.fr",
        "Lien_Depot_AAP": "https://www.fondationbanquepopulaire.fr/deposer-un-projet/",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Type_Approche": "Email direct + Dépôt de candidature",
        "Angle_Pitch_Ambition_Campus": "Révéler l'audace et briser le plafond de verre des lycéens issus des quartiers.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Prendre contact avec l'équipe mécénat à Paris 13e."
    },
    {
        "ID": "FOND-10",
        "Nom_Fondation": "Fonds AXA pour le Progrès Humain & AXA Atout Cœur",
        "Groupe_Parent": "AXA France",
        "Priorite": "Tier 2 - Grand Donateur",
        "Thematiques_Cibles": "Inclusion sociale, Éducation, Égalité des chances QPV",
        "Nom_Contact": "Direction Mécénat AXA France",
        "Poste_Contact": "Responsable Partenariats Associatifs",
        "Email_Contact": "axa.atoutcoeur@axa.fr / via axa.fr",
        "Site_Web": "https://www.axa.com/fr/le-fonds-axa-pour-le-progres-humain",
        "Lien_Depot_AAP": "Plateforme partenariats associatifs AXA",
        "Ticket_Moyen_Estime": "10 000 € - 25 000 €",
        "Type_Approche": "Mécénat financier + Mobilisation de mentors AXA",
        "Angle_Pitch_Ambition_Campus": "Coût de 35 €/jeune et impact mesuré : ratio 1€ investi = 5,30€ de valeur terrain.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Proposer une convention de soutien financier couplée au bénévolat AXA Atout Cœur."
    },
    {
        "ID": "FOND-11",
        "Nom_Fondation": "Fondation TotalEnergies",
        "Groupe_Parent": "TotalEnergies SE",
        "Priorite": "Tier 2 - Grand Donateur",
        "Thematiques_Cibles": "Éducation, Insertion des jeunes vulnérables, Égalité des chances",
        "Nom_Contact": "Jacques-Emmanuel Saulnier",
        "Poste_Contact": "Délégué Général Fondation TotalEnergies",
        "Email_Contact": "fondation@totalenergies.com / 01 47 44 45 46",
        "Site_Web": "https://fondation.totalenergies.com",
        "Lien_Depot_AAP": "https://fondation.totalenergies.com/fr/candidater (Plateforme annuelle)",
        "Ticket_Moyen_Estime": "20 000 € - 50 000 €",
        "Type_Approche": "Candidature AAP national / territorial + Contact institutionnel",
        "Angle_Pitch_Ambition_Campus": "Grande échelle : 500+ lycéens/an, 36 lycées conventionnés sur 4 régions françaises.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Surveiller l'ouverture du prochain AAP national ou déposer une demande spontanée."
    },
    {
        "ID": "FOND-12",
        "Nom_Fondation": "Fondation Groupe SNCF",
        "Groupe_Parent": "SNCF",
        "Priorite": "Tier 3 - Partenaire Institutionnel",
        "Thematiques_Cibles": "Insertion des jeunes, Réussite éducative, Construction de parcours de vie",
        "Nom_Contact": "Laetitia Gourbeille",
        "Poste_Contact": "Déléguée Générale Fondation Groupe SNCF",
        "Email_Contact": "fondation@sncf.fr",
        "Site_Web": "https://www.fondation-sncf.org",
        "Lien_Depot_AAP": "Dispositif Coups de Cœur Citoyens & Partenariats pluriannuels",
        "Ticket_Moyen_Estime": "10 000 € - 25 000 €",
        "Type_Approche": "Email direct de prise de contact institutionnelle",
        "Angle_Pitch_Ambition_Campus": "Accompagnement dans la durée (mentorat sur 1 à 3 ans) et présence territoriale (Paris, Reims, Poitiers, Menton).",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Envoyer une présentation synthétique à fondation@sncf.fr."
    },
    {
        "ID": "FOND-13",
        "Nom_Fondation": "Fondation Groupe RATP",
        "Groupe_Parent": "Groupe RATP",
        "Priorite": "Tier 3 - Partenaire Institutionnel",
        "Thematiques_Cibles": "Insertion sociale, Éducation, Égalité des chances en Île-de-France",
        "Nom_Contact": "Direction de la Fondation Groupe RATP",
        "Poste_Contact": "Secrétaire Général & Responsables Mécénat",
        "Email_Contact": "fondation@ratp.fr",
        "Site_Web": "https://www.ratp.fr/groupe-ratp/fondation-groupe-ratp",
        "Lien_Depot_AAP": "Appels à projets thématiques annuels Île-de-France",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Type_Approche": "Candidature AAP + Approche territoriale francilienne",
        "Angle_Pitch_Ambition_Campus": "Forte implantation francilienne (445 binômes actifs dans les départements 75, 92, 93, 94).",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Positionner le dossier sur la mobilité sociale et l'accès aux études supérieures."
    },
    {
        "ID": "FOND-14",
        "Nom_Fondation": "Fondation Groupe ADP (Aéroports de Paris)",
        "Groupe_Parent": "Groupe ADP",
        "Priorite": "Tier 3 - Partenaire Institutionnel",
        "Thematiques_Cibles": "Éducation, Prévention du décrochage, Insertion jeunes en Île-de-France",
        "Nom_Contact": "Direction RSE & Mécénat Groupe ADP",
        "Poste_Contact": "Responsable des Partenariats Mécénat",
        "Email_Contact": "fondation@adp.fr",
        "Site_Web": "https://www.parisaeroport.fr",
        "Lien_Depot_AAP": "Appels à projets annuels (Ouverture Janvier-Mars)",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Type_Approche": "Candidature AAP sur les territoires 93/94/95 proches aéroports",
        "Angle_Pitch_Ambition_Campus": "Accompagnement intensif des lycées situés sur les bassins Roissy / Orly.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Cibler les lycées conventionnés situés à proximité des plateformes aéroportuaires."
    },
    {
        "ID": "FOND-15",
        "Nom_Fondation": "Fondation Orange",
        "Groupe_Parent": "Orange",
        "Priorite": "Tier 3 - Télécoms & Tech",
        "Thematiques_Cibles": "Éducation numérique, Insertion des jeunes et des femmes, Tiers-Lieux Solidaires",
        "Nom_Contact": "Direction de la Fondation Orange",
        "Poste_Contact": "Responsable Programmes Éducation & Insertion",
        "Email_Contact": "fondation.orange@orange.com",
        "Site_Web": "https://www.fondationorange.com",
        "Lien_Depot_AAP": "https://www.fondationorange.com/fr/candidater (Plateforme AAP)",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Type_Approche": "Candidature AAP + Mécénat de compétences",
        "Angle_Pitch_Ambition_Campus": "Digitalisation des oraux blancs et accès aux outils d'orientation pour lycéens REP.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Contacter l'antenne régionale Île-de-France à Issy-les-Moulineaux (111 quai Roosevelt)."
    },
    {
        "ID": "FOND-16",
        "Nom_Fondation": "Fondation Free (Groupe Iliad)",
        "Groupe_Parent": "Iliad / Free",
        "Priorite": "Tier 3 - Télécoms & Tech",
        "Thematiques_Cibles": "Inclusion numérique, Émancipation, Égalité des chances par l'éducation",
        "Nom_Contact": "Direction de la Fondation Free",
        "Poste_Contact": "Responsable Mécénat & Projets",
        "Email_Contact": "fondation@iliad.fr",
        "Site_Web": "https://www.iliad.fr/fr/engagements/fondation-free",
        "Lien_Depot_AAP": "Appels à projets réguliers sur la plateforme Iliad",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Type_Approche": "Dépôt AAP + Mise en avant de l'audace et de l'anticonformisme",
        "Angle_Pitch_Ambition_Campus": "Briser les codes : « Rendre la pareille, c'est notre identité », modèle 100% bénévole.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Valoriser le lien avec l'esprit d'émancipation de l'École 42 / Station F (lieux de visite)."
    },
    {
        "ID": "FOND-17",
        "Nom_Fondation": "Fondation ENGIE",
        "Groupe_Parent": "ENGIE",
        "Priorite": "Tier 3 - Énergie & Climat",
        "Thematiques_Cibles": "Inclusion de la jeunesse, Éducation, Égalité des chances, Emploi",
        "Nom_Contact": "Direction Mécénat ENGIE",
        "Poste_Contact": "Responsable des Projets Jeunesse",
        "Email_Contact": "fondation.engie@engie.com",
        "Site_Web": "https://projets.fondation-engie.com",
        "Lien_Depot_AAP": "https://projets.fondation-engie.com/ (Plateforme de dépôt en continu)",
        "Ticket_Moyen_Estime": "10 000 € - 25 000 €",
        "Type_Approche": "Dépôt direct sur portail officiel",
        "Angle_Pitch_Ambition_Campus": "Impact mesurable et systémique : 17 ans de recul, note de satisfaction 9,2/10.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Créer un compte sur le portail projets.fondation-engie.com et soumettre le projet."
    },
    {
        "ID": "FOND-18",
        "Nom_Fondation": "Fondation Francis Bouygues",
        "Groupe_Parent": "Groupe Bouygues",
        "Priorite": "Tier 3 - Grand Groupe",
        "Thematiques_Cibles": "Bourses d'excellence, Parrainage de bacheliers méritants de milieux modestes",
        "Nom_Contact": "Direction de la Fondation Francis Bouygues",
        "Poste_Contact": "Délégué Général & Responsable Bourses",
        "Email_Contact": "fondationfrancisbouygues@bouygues.com",
        "Site_Web": "https://www.fondationfrancisbouygues.com",
        "Lien_Depot_AAP": "Contact direct partenariats associatifs",
        "Ticket_Moyen_Estime": "10 000 € - 25 000 €",
        "Type_Approche": "Partenariat de sourcing bourses + Soutien financier",
        "Angle_Pitch_Ambition_Campus": "Ambition Campus prépare et coache les bacheliers qui intègrent les filières d'excellence (Sciences Po, CPGE).",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Proposer un partenariat passerelle pour flécher les lycéens admis vers leurs bourses d'études."
    },
    {
        "ID": "FOND-19",
        "Nom_Fondation": "Fondation TF1",
        "Groupe_Parent": "Groupe TF1 (Bouygues)",
        "Priorite": "Tier 3 - Médias",
        "Thematiques_Cibles": "Égalité des chances, Diversité dans les médias, Prise de parole & Éloquence",
        "Nom_Contact": "Direction RSE & Fondation TF1",
        "Poste_Contact": "Responsable Fondation TF1",
        "Email_Contact": "fondationtf1@tf1.fr",
        "Site_Web": "https://groupe-tf1.fr/fr/engagements-rse/notre-fondation",
        "Lien_Depot_AAP": "Contact direct partenariats",
        "Ticket_Moyen_Estime": "8 000 € - 15 000 €",
        "Type_Approche": "Mécénat financier + Parrainage du concours annuel d'éloquence",
        "Angle_Pitch_Ambition_Campus": "Ateliers d'art oratoire (Ethos/Pathos/Logos), AC Décrypte et documentaire « Mérite sous condition ».",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Proposer à TF1 de parrainer la finale du concours d'éloquence et de composer le jury."
    },
    {
        "ID": "FOND-20",
        "Nom_Fondation": "Fondation FDJ (FDJ United)",
        "Groupe_Parent": "La Française des Jeux (FDJ)",
        "Priorite": "Tier 3 - Grand Groupe",
        "Thematiques_Cibles": "Égalité des chances par le jeu et l'éducation, Insertion des publics vulnérables",
        "Nom_Contact": "Direction de la Fondation FDJ",
        "Poste_Contact": "Responsable Mécénat & Appels à Projets",
        "Email_Contact": "fondation@fdjunited.com",
        "Site_Web": "https://www.fdjunited.com",
        "Lien_Depot_AAP": "https://www.fdjunited.com (Appels à projets thématiques annuels)",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Type_Approche": "Candidature AAP + Email d'accroche",
        "Angle_Pitch_Ambition_Campus": "Pédagogie active et émancipation : ateliers oraux blancs ludifiés, simulations d'entretiens.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Surveiller les dates du prochain appel à projets Égalité des chances FDJ."
    },
    {
        "ID": "FOND-21",
        "Nom_Fondation": "Fondation AlphaOmega",
        "Groupe_Parent": "Venture Philanthropy France",
        "Priorite": "Tier 4 - Venture Philanthropy",
        "Thematiques_Cibles": "Réussite scolaire, Lutte contre le décrochage, Égalité des chances (Changement d'échelle)",
        "Nom_Contact": "Équipe Investissement & Partenariats",
        "Poste_Contact": "Directeur des Investissements Philanthropiques",
        "Email_Contact": "contact@alphaomega-fondation.org",
        "Site_Web": "https://www.alphaomega-fondation.org",
        "Lien_Depot_AAP": "Candidature partenariats pluriannuels",
        "Ticket_Moyen_Estime": "25 000 € - 75 000 €",
        "Type_Approche": "Dossier stratégique de changement d'échelle",
        "Angle_Pitch_Ambition_Campus": "Modèle d'accompagnement éprouvé (17 ans), 525 binômes actifs, fort potentiel d'expansion nationale.",
        "Statut_Prospection": "À préparer pour échange stratégique",
        "Notes_Action": "Préparer un pitch d'expansion nationale sur le modèle du Social ROI."
    },
    {
        "ID": "FOND-22",
        "Nom_Fondation": "Break Poverty Foundation",
        "Groupe_Parent": "Fondation Reconnue d'Utilité Publique",
        "Priorite": "Tier 4 - Philanthropie d'Impact",
        "Thematiques_Cibles": "Prévention de la pauvreté, Mentorat, Accompagnement à l'orientation et à l'insertion",
        "Nom_Contact": "Équipe Projets & Partenariats",
        "Poste_Contact": "Responsable des Programmes Territoires Zéro Décrochage",
        "Email_Contact": "contact@breakpoverty.com (01 85 76 50 50)",
        "Site_Web": "https://www.breakpoverty.com",
        "Lien_Depot_AAP": "Programmes territoriaux d'insertion",
        "Ticket_Moyen_Estime": "15 000 € - 35 000 €",
        "Type_Approche": "Prise de contact institutionnelle (Siège : 81 rue de Lille, Paris 7e)",
        "Angle_Pitch_Ambition_Campus": "Mentorat individuel de 500+ jeunes en milieu populaire comme bouclier contre l'autocensure.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Contacter l'équipe partenariats par email et téléphone."
    },
    {
        "ID": "FOND-23",
        "Nom_Fondation": "Fondation FACE (Agir Contre l'Exclusion)",
        "Groupe_Parent": "Fondation Reconnue d'Utilité Publique",
        "Priorite": "Tier 4 - Réseau Entreprises & QPV",
        "Thematiques_Cibles": "Lutte contre l'exclusion, Égalité des chances, Insertion en QPV",
        "Nom_Contact": "Direction des Partenariats Nationaux",
        "Poste_Contact": "Responsable Projets Éducation & Territoires",
        "Email_Contact": "contact@fondationface.org",
        "Site_Web": "https://www.fondationface.org",
        "Lien_Depot_AAP": "Concours « S'engager pour les Quartiers » & Appels à projets",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Type_Approche": "Candidature concours quartiers + conventionnement réseau",
        "Angle_Pitch_Ambition_Campus": "Impact direct dans les quartiers prioritaires : 36 lycées REP, 75 mentors engagés.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Candidater au concours annuel « S'engager pour les Quartiers » de la Fondation FACE."
    },
    {
        "ID": "FOND-24",
        "Nom_Fondation": "Fondation de France (Programme Éducation)",
        "Groupe_Parent": "1er Réseau Philanthropique de France",
        "Priorite": "Tier 4 - Réseau Fondations",
        "Thematiques_Cibles": "Éducation, Enfance, Émancipation des jeunes, Lutte contre le déterminisme",
        "Nom_Contact": "Département Philanthropie & Mécénat",
        "Poste_Contact": "Responsable du Programme Enfance / Éducation",
        "Email_Contact": "donateurs@fdf.org",
        "Site_Web": "https://www.fondationdefrance.org",
        "Lien_Depot_AAP": "https://www.fondationdefrance.org/fr/trouver-un-financement",
        "Ticket_Moyen_Estime": "10 000 € - 30 000 €",
        "Type_Approche": "Candidature aux AAP thématiques Éducation & Jeunesse",
        "Angle_Pitch_Ambition_Campus": "Action continue sur le terrain depuis 2008, 100% bénévole, transparence financière totale.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Identifier l'AAP annuel « Éducation » et soumettre le projet de rentrée."
    },
    {
        "ID": "FOND-25",
        "Nom_Fondation": "Fondation L'Oréal",
        "Groupe_Parent": "L'Oréal",
        "Priorite": "Tier 4 - Grand Donateur",
        "Thematiques_Cibles": "Émancipation, Accès à l'éducation supérieure, Bourses d'excellence",
        "Nom_Contact": "Direction de la Fondation L'Oréal",
        "Poste_Contact": "Responsable Programmes Éducation & Inclusion",
        "Email_Contact": "fondation.loreal@loreal.com (01 47 56 70 00)",
        "Site_Web": "https://www.fondationloreal.com",
        "Lien_Depot_AAP": "Contact direct partenariats d'émancipation",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Type_Approche": "Email d'accroche direct sur l'émancipation des lycéennes de QPV",
        "Angle_Pitch_Ambition_Campus": "Forte proportion de lycéennes accompagnées accédant aux prépas et grandes écoles (Sciences Po, Assas, Sorbonne).",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Adresser une proposition de parrainage de promotion à fondation.loreal@loreal.com."
    },
    {
        "ID": "FOND-26",
        "Nom_Fondation": "Fondation Decathlon",
        "Groupe_Parent": "Decathlon",
        "Priorite": "Tier 4 - Grand Donateur",
        "Thematiques_Cibles": "Inclusion sociale, Éducation, Émancipation des jeunes vulnérables",
        "Nom_Contact": "Marie Pinel / Thomas Dumortier",
        "Poste_Contact": "Leader France / Chef de projet impact social",
        "Email_Contact": "marie.pinel@decathlon.com / thomas.dumortier@decathlon.com",
        "Site_Web": "https://www.fondationdecathlon.com",
        "Lien_Depot_AAP": "Candidature partenariats locaux et nationaux",
        "Ticket_Moyen_Estime": "8 000 € - 18 000 €",
        "Type_Approche": "Email direct aux responsables d'impact social",
        "Angle_Pitch_Ambition_Campus": "Posture, confiance en soi et dépassement personnel à travers l'éloquence et les concours.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Envoyer un email d'approche à Marie Pinel et Thomas Dumortier."
    }
]


def clean_text_field(text):
    """Nettoie les artefacts de recherche web (citations +1, balises, etc.)."""
    if not text or pd.isna(text):
        return ""
    t = str(text)
    # Nettoyer les balises de citations web (+1, +2, etc.)
    t = re.sub(r'(\w+[\.\-\w]*)\+\d+', r'\1', t)
    # Nettoyer les suffixes d'URL collés
    t = re.sub(r'(https?://[^\s]+)(fondation\-[a-z]+|eiffage|vinci|pressroom\.[a-z]+|carenews|laposte[a-z]+|corporate\.[a-z\.]+)', r'\1', t)
    # Espaces multiples
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def load_and_merge_foundations():
    """Charge le fichier téléchargé, le nettoie et le fusionne avec le lot initial."""
    all_data = list(BASE_FONDATIONS_LOT1)

    if os.path.exists(DOWNLOAD_PATH):
        print(f"Chargement des nouvelles données depuis : {DOWNLOAD_PATH}")
        with open(DOWNLOAD_PATH, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.DictReader(f)
            for row in reader:
                fid = clean_text_field(row.get("ID", ""))
                # Ne pas dupliquer si déjà présent
                if any(x["ID"] == fid for x in all_data):
                    continue

                nom = clean_text_field(row.get("Nom_Fondation", ""))
                parent = clean_text_field(row.get("Groupe_Parent", ""))
                theme = clean_text_field(row.get("Thematiques_Cibles", ""))
                contact = clean_text_field(row.get("Nom_Contact", ""))
                poste = clean_text_field(row.get("Poste_Contact", ""))
                email = clean_text_field(row.get("Email_Contact_Verifie", ""))
                web = clean_text_field(row.get("Site_Web_Fondation", ""))
                aap = clean_text_field(row.get("Lien_Depot_AAP", ""))
                ticket = clean_text_field(row.get("Ticket_Moyen_Estime", ""))
                pitch = clean_text_field(row.get("Angle_Pitch_Ambition_Campus", ""))

                # Nettoyage spécifique sur les emails
                if "pattern probable" in email or not email:
                    email = f"contact via {web} (direction mécénat)"

                # Attribution de la priorité selon le profil
                if any(k in nom.lower() for k in ["vinci", "eiffage", "nexity", "saint-gobain", "gecina", "crédit mutuel", "crédit agricole", "caisse d’épargne", "banque populaire"]):
                    prio = "Tier 2 - Grand Donateur / BTP & Banque"
                elif any(k in nom.lower() for k in ["sopra steria", "devoteam", "capgemini", "la poste", "air france"]):
                    prio = "Tier 3 - Tech, Transport & Médias"
                else:
                    prio = "Tier 4 - Fondation Abritée & Enseignement Supérieur"

                item = {
                    "ID": fid,
                    "Nom_Fondation": nom,
                    "Groupe_Parent": parent,
                    "Priorite": prio,
                    "Thematiques_Cibles": theme,
                    "Nom_Contact": contact,
                    "Poste_Contact": poste,
                    "Email_Contact": email,
                    "Site_Web": web,
                    "Lien_Depot_AAP": aap,
                    "Ticket_Moyen_Estime": ticket,
                    "Type_Approche": "Candidature AAP / Contact mécénat direct",
                    "Angle_Pitch_Ambition_Campus": pitch,
                    "Statut_Prospection": "À qualifier & contacter",
                    "Notes_Action": f"Vérifier la session AAP sur {web} et envoyer le PDF A4."
                }
                all_data.append(item)

    return all_data


def generate_database():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    foundations = load_and_merge_foundations()
    df = pd.DataFrame(foundations)

    # 1. Export CSV propre encodé UTF-8 BOM
    df.to_csv(CSV_PATH, index=False, encoding="utf-8-sig", sep=";")
    print(f"[OK] Master CSV Fondations genere : {CSV_PATH} ({len(df)} fondations au total)")

    # 2. Export Excel formaté professionnellement
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Master_Fondations", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Master_Fondations"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F1E36", end_color="0F1E36", fill_type="solid")
        cell_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_widths = {
            "ID": 10,
            "Nom_Fondation": 30,
            "Groupe_Parent": 26,
            "Priorite": 22,
            "Thematiques_Cibles": 32,
            "Nom_Contact": 26,
            "Poste_Contact": 30,
            "Email_Contact": 32,
            "Site_Web": 30,
            "Lien_Depot_AAP": 35,
            "Ticket_Moyen_Estime": 20,
            "Type_Approche": 32,
            "Angle_Pitch_Ambition_Campus": 38,
            "Statut_Prospection": 24,
            "Notes_Action": 35,
        }

        for row_idx in range(2, len(df) + 2):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if col_name == "Priorite":
                    val = str(cell.value)
                    if "Tier 1" in val:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="166534")
                    elif "Tier 2" in val:
                        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="1E40AF")
                    elif "Tier 3" in val:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")

        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = col_widths.get(col_name, 22)

        worksheet.freeze_panes = "A2"

    print(f"[OK] Master Excel Fondations genere : {XLSX_PATH}")


if __name__ == "__main__":
    generate_database()
