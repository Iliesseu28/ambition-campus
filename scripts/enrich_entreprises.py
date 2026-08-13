"""
Ambition Campus — Base de Données Master & Enrichissement Mécénat Privé d'Entreprise
Pilier 2 : 48 Entreprises Cibles de Mécénat Direct en France.
Génère le CSV complet et le fichier Excel formaté avec tableau de bord et filtres sectoriels.
"""

import os
import re
import csv
import pandas as pd

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "..", "prospection", "entreprises")
CSV_PATH = os.path.join(OUTPUT_DIR, "entreprises_database.csv")
XLSX_PATH = os.path.join(OUTPUT_DIR, "entreprises_database.xlsx")
DOWNLOAD_PATH = r"C:\Users\User\Downloads\ID-NomEntreprise-SecteurActivite-NomContactDecideu.csv"

# ── LOT 1 : 26 PREMIÈRES ENTREPRISES QUALIFIÉES (TIER 1 À TIER 3) ──
BASE_ENTREPRISES_LOT1 = [
    {
        "ID": "ENT-01",
        "Nom_Entreprise": "PwC France et Maghreb",
        "Secteur_Activite": "Audit & Conseil",
        "Priorite": "Tier 1 - Partenaire Historique",
        "Nom_Contact": "Adélaïde de Tourtier / Candice Galopeau",
        "Poste_Contact": "Directrice RSE & Engagement / Resp. Engagement Sociétal",
        "Email_Contact": "adelaide.de.tourtier@pwc.com / candice.galopeau@pwc.com",
        "LinkedIn_Contact": "https://www.linkedin.com/in/adelaide-de-tourtier",
        "Site_Web": "https://www.pwc.fr/fr/fondation.html",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Upgrade Partenariat Jurys -> Mécénat Financier Annuel",
        "Angle_Pitch_Ambition_Campus": "Vos collaborateurs sont déjà jurys aux oraux blancs. Devenez Mécène Partenaire Officiel 2026-2027.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Proposer une convention annuelle de mécénat avec masterclass dans leurs locaux de Neuilly."
    },
    {
        "ID": "ENT-02",
        "Nom_Entreprise": "Deloitte France",
        "Secteur_Activite": "Audit & Conseil",
        "Priorite": "Tier 1 - Partenaire Historique",
        "Nom_Contact": "Pôle Attractivité & RSE",
        "Poste_Contact": "Responsable Mécénat & Diversité",
        "Email_Contact": "frpoleattractivite@deloitte.fr",
        "LinkedIn_Contact": "https://www.linkedin.com/company/deloitte",
        "Site_Web": "https://www2.deloitte.com/fr/fr.html",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Upgrade Partenariat Jurys -> Convention Mécénat",
        "Angle_Pitch_Ambition_Campus": "Pérennisation du pôle des 210+ oraux blancs et parrainage d'une promotion de 50 lycéens REP.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Cibler la Tour Majunga (Paris La Défense)."
    },
    {
        "ID": "ENT-03",
        "Nom_Entreprise": "KPMG France",
        "Secteur_Activite": "Audit & Conseil (Entreprise à Mission)",
        "Priorite": "Tier 1 - Partenaire Historique",
        "Nom_Contact": "Direction Engagement Citoyen",
        "Poste_Contact": "Directeur Engagement Citoyen & Inclusion",
        "Email_Contact": "contact via Tour Eqho Courbevoie (prenom.nom@kpmg.fr)",
        "LinkedIn_Contact": "https://www.linkedin.com/company/kpmg-france",
        "Site_Web": "https://kpmg.com/fr/fr/about/engagement-citoyen.html",
        "Ticket_Moyen_Estime": "10 000 € - 15 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Synergie programme « Les Lycées de la Réussite »",
        "Angle_Pitch_Ambition_Campus": "Alignement parfait entre vos 36 lycées REP conventionnés et leur stratégie d'entreprise à mission.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Solliciter un échange de 15 minutes avec l'équipe Engagement Citoyen."
    },
    {
        "ID": "ENT-04",
        "Nom_Entreprise": "EY France (Ernst & Young)",
        "Secteur_Activite": "Audit & Conseil",
        "Priorite": "Tier 1 - Partenaire Historique",
        "Nom_Contact": "Fabienne Marqueste / Orane Tribouley",
        "Poste_Contact": "Déléguée Générale Mécénat / Resp. accompagnement projets",
        "Email_Contact": "fabienne.marqueste@fr.ey.com / orane.tribouley@fr.ey.com",
        "LinkedIn_Contact": "https://www.linkedin.com/in/fabienne-marqueste",
        "Site_Web": "https://www.ey.com/fr_fr",
        "Ticket_Moyen_Estime": "8 000 € - 15 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Upgrade Partenariat Jurys & Mécénat de Compétences",
        "Angle_Pitch_Ambition_Campus": "Co-financement des stages intensifs d'éloquence et des bourses de mobilité étudiante.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Contacter Orane Tribouley à la Tour First (Paris La Défense)."
    },
    {
        "ID": "ENT-05",
        "Nom_Entreprise": "Forvis Mazars France",
        "Secteur_Activite": "Audit, Fiscalité & Conseil",
        "Priorite": "Tier 2 - Grand Cabinet",
        "Nom_Contact": "Marie-Anne Brin",
        "Poste_Contact": "Directrice RSE & Fondation Forvis Mazars",
        "Email_Contact": "marie-anne.brin@mazars.fr",
        "LinkedIn_Contact": "https://www.linkedin.com/in/marie-anne-brin",
        "Site_Web": "https://www.mazars.fr",
        "Ticket_Moyen_Estime": "8 000 € - 15 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Sollicitation Mécénat Égalité des Chances & Diversité",
        "Angle_Pitch_Ambition_Campus": "Sourcing de profils d'excellence issus de QPV pour leurs filières d'audit et de conseil.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Envoyer le template d'email RSE au siège Tour Exaltis La Défense."
    },
    {
        "ID": "ENT-06",
        "Nom_Entreprise": "Wavestone",
        "Secteur_Activite": "Conseil en Management & SI",
        "Priorite": "Tier 2 - Grand Cabinet",
        "Nom_Contact": "Direction RSE & Mécénat",
        "Poste_Contact": "Responsable RSE & Partenariats Solidaires",
        "Email_Contact": "contact@wavestone.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/wavestone",
        "Site_Web": "https://www.wavestone.com/fr/",
        "Ticket_Moyen_Estime": "5 000 € - 12 000 €",
        "Levier_Fiscal_60pct": "Don 8 000 € = Coût net 3 200 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat financier + Bénévolat de compétences",
        "Angle_Pitch_Ambition_Campus": "Accompagnement de lycéens vers les écoles d'ingénieurs et filières scientifiques/numériques.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Cibler l'équipe RSE à la Tour Franklin (La Défense)."
    },
    {
        "ID": "ENT-07",
        "Nom_Entreprise": "Boston Consulting Group (BCG Paris)",
        "Secteur_Activite": "Conseil en Stratégie",
        "Priorite": "Tier 2 - Cabinet de Stratégie",
        "Nom_Contact": "Pôle Social Impact & Diversity",
        "Poste_Contact": "Responsable Social Impact & Égalité des Chances",
        "Email_Contact": "bcg.paris@bcg.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/boston-consulting-group",
        "Site_Web": "https://www.bcg.com/fr-fr",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Levier_Fiscal_60pct": "Don 20 000 € = Coût net 8 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Partenariat Social Impact & Soutien Financier Direct",
        "Angle_Pitch_Ambition_Campus": "Programme d'élite brisant l'autocensure vers Sciences Po, Sorbonne, Prépas (Henri IV, Saint-Louis).",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Adresser une note synthétique à l'équipe Social Impact au 75 av de la Grande Armée (Paris 16e)."
    },
    {
        "ID": "ENT-08",
        "Nom_Entreprise": "McKinsey & Company Paris",
        "Secteur_Activite": "Conseil en Stratégie",
        "Priorite": "Tier 2 - Cabinet de Stratégie",
        "Nom_Contact": "Alain Imbert / Christophe Rohel",
        "Poste_Contact": "Responsable Appels à Projets & Mécénat / Resp. Recrutement",
        "Email_Contact": "fro_projet@mckinsey.com / fro_recrutement@mckinsey.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/mckinsey",
        "Site_Web": "https://www.mckinsey.com/fr",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Levier_Fiscal_60pct": "Don 20 000 € = Coût net 8 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Email direct au pôle Projets & Mécénat",
        "Angle_Pitch_Ambition_Campus": "Réseau d'égalité des chances à fort Social ROI (1€=5,30€), 21 admis Sciences Po 2026.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Envoyer le pitch à fro_projet@mckinsey.com (35 bd des Invalides, Paris 7e)."
    },
    {
        "ID": "ENT-09",
        "Nom_Entreprise": "Banque de France",
        "Secteur_Activite": "Institution Financière & Banque Centrale",
        "Priorite": "Tier 1 - Partenaire Historique",
        "Nom_Contact": "Direction RSE & Mécénat",
        "Poste_Contact": "Responsable Mécénat Participatif & Inclusion",
        "Email_Contact": "rse@banque-france.fr / via banque-france.fr",
        "LinkedIn_Contact": "https://www.linkedin.com/company/banque-de-france",
        "Site_Web": "https://www.banque-france.fr",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Convention de soutien institutionnel & mécénat",
        "Type_Approche": "Upgrade Partenariat Visites/Jurys -> Mécénat Annuel",
        "Angle_Pitch_Ambition_Campus": "Collaborateurs déjà impliqués. Inclusion économique et parrainage des jeunes de milieux populaires.",
        "Statut_Prospection": "À contacter en priorité",
        "Notes_Action": "Proposer d'inscrire Ambition Campus au dispositif de mécénat collaborateur « Vos voix, nos dons »."
    },
    {
        "ID": "ENT-10",
        "Nom_Entreprise": "BNP Paribas (Direction RSE / RH)",
        "Secteur_Activite": "Banque & Services Financiers",
        "Priorite": "Tier 2 - Grand Groupe Bancaire",
        "Nom_Contact": "Direction de l'Engagement d'Entreprise",
        "Poste_Contact": "Responsable RSE & Relations Écoles France",
        "Email_Contact": "engagement.entreprise@bnpparibas.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/bnp-paribas",
        "Site_Web": "https://group.bnpparibas",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Levier_Fiscal_60pct": "Don 15 000 € = Coût net 6 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat financier direct d'entreprise (complémentaire à la fondation)",
        "Angle_Pitch_Ambition_Campus": "Soutien direct aux lycéens de banlieue vers les filières financières, juridiques et économiques.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Contacter la Direction de l'Engagement au siège bd des Italiens (Paris 9e)."
    },
    {
        "ID": "ENT-11",
        "Nom_Entreprise": "Crédit Agricole CIB (Corporate & Investment Bank)",
        "Secteur_Activite": "Banque de Financement et d'Investissement",
        "Priorite": "Tier 2 - BFI & Marchés",
        "Nom_Contact": "Direction RSE & Diversité CACIB",
        "Poste_Contact": "Responsable Inclusion & Mécénat BFI",
        "Email_Contact": "rse@ca-cib.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/credit-agricole-cib",
        "Site_Web": "https://www.ca-cib.com",
        "Ticket_Moyen_Estime": "10 000 € - 25 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat Diversité Sociale & Bourses d'excellence",
        "Angle_Pitch_Ambition_Campus": "Ouverture des métiers de la finance aux profils talentueux issus de la diversité.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Cibler le campus Evergreen à Montrouge (92)."
    },
    {
        "ID": "ENT-12",
        "Nom_Entreprise": "Groupe BPCE / Natixis",
        "Secteur_Activite": "Banque & Gestion d'Actifs",
        "Priorite": "Tier 2 - BFI & Gestion d'Actifs",
        "Nom_Contact": "Direction RSE Groupe BPCE",
        "Poste_Contact": "Directeur RSE & Mécénat Territorial",
        "Email_Contact": "rse@bpce.fr",
        "LinkedIn_Contact": "https://www.linkedin.com/company/groupe-bpce",
        "Site_Web": "https://groupebpce.com",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Convention de parrainage de promotions régionales",
        "Angle_Pitch_Ambition_Campus": "Couverture sur Paris/IDF, Reims, Poitiers et Menton correspondant au maillage régional BPCE.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Contacter la direction RSE quai d'Austerlitz (Paris 13e)."
    },
    {
        "ID": "ENT-13",
        "Nom_Entreprise": "Lazard Frères SAS (Paris)",
        "Secteur_Activite": "Banque d'Affaires & Conseil M&A",
        "Priorite": "Tier 2 - Banque d'Affaires d'Élite",
        "Nom_Contact": "Secrétariat Général & RSE Lazard Paris",
        "Poste_Contact": "Secrétaire Général / Responsable RSE",
        "Email_Contact": "contact via lazard.com (175 bd Haussmann Paris)",
        "LinkedIn_Contact": "https://www.linkedin.com/company/lazard",
        "Site_Web": "https://www.lazard.com/fr",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Levier_Fiscal_60pct": "Don 15 000 € = Coût net 6 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat d'excellence & Bourses d'études",
        "Angle_Pitch_Ambition_Campus": "Propulser les bacheliers brillants de banlieue vers les grandes écoles (Dauphine, Sciences Po, ESSEC).",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Adresser un courrier de proposition de partenariat mécénat au 175 boulevard Haussmann."
    },
    {
        "ID": "ENT-14",
        "Nom_Entreprise": "Rothschild & Co (Paris)",
        "Secteur_Activite": "Banque d'Affaires & Banque Privée",
        "Priorite": "Tier 2 - Banque d'Affaires d'Élite",
        "Nom_Contact": "Direction Sustainability & Community Affairs",
        "Poste_Contact": "Responsable Mécénat & Impact Social",
        "Email_Contact": "via rothschildandco.com (avenue de Messine Paris)",
        "LinkedIn_Contact": "https://www.linkedin.com/company/rothschild-co",
        "Site_Web": "https://www.rothschildandco.com/fr",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Levier_Fiscal_60pct": "Don 15 000 € = Coût net 6 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat d'impact social & Égalité des chances",
        "Angle_Pitch_Ambition_Campus": "Accompagnement d'excellence dans la durée, taux de réussite exceptionnel aux concours sélectifs.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Cibler le département Sustainability au 23 bis avenue de Messine (Paris 8e)."
    },
    {
        "ID": "ENT-15",
        "Nom_Entreprise": "Gide Loyrette Nouel",
        "Secteur_Activite": "Cabinet d'Avocats d'Affaires",
        "Priorite": "Tier 2 - Droit & Juridique",
        "Nom_Contact": "Comité Gide Pro Bono & RSE",
        "Poste_Contact": "Associé référent Pro Bono & Bourses",
        "Email_Contact": "probono@gide.com / via gide.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/gide-loyrette-nouel",
        "Site_Web": "https://www.gide.com",
        "Ticket_Moyen_Estime": "10 000 € - 25 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat financier + Bourses d'excellence Droit (Sorbonne, Assas, Sciences Po)",
        "Angle_Pitch_Ambition_Campus": "17 admis à La Sorbonne et filières juridiques sélectives formés par Ambition Campus.",
        "Statut_Prospection": "À contacter en priorité (Secteur Droit)",
        "Notes_Action": "Prendre contact avec le pôle Pro Bono au 15 rue de Laborde (Paris 8e)."
    },
    {
        "ID": "ENT-16",
        "Nom_Entreprise": "August Debouzy",
        "Secteur_Activite": "Cabinet d'Avocats d'Affaires",
        "Priorite": "Tier 2 - Droit & Juridique",
        "Nom_Contact": "Direction RSE & Pro Bono",
        "Poste_Contact": "Responsable RSE & Engagements Solidaires",
        "Email_Contact": "contact@august-debouzy.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/august-&-debouzy",
        "Site_Web": "https://www.august-debouzy.com",
        "Ticket_Moyen_Estime": "8 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat financier + Jurys d'éloquence et prise de parole",
        "Angle_Pitch_Ambition_Campus": "Concours annuel d'éloquence et ateliers oratoires animés avec des avocats du cabinet.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Proposer aux avocats d'August Debouzy d'être jurys de la finale d'éloquence."
    },
    {
        "ID": "ENT-17",
        "Nom_Entreprise": "Clifford Chance Paris",
        "Secteur_Activite": "Cabinet d'Avocats International",
        "Priorite": "Tier 2 - Droit & Juridique",
        "Nom_Contact": "Responsible Business & Inclusion Committee",
        "Poste_Contact": "Responsable Inclusion & Pro Bono Paris",
        "Email_Contact": "via 1 rue d'Astorg Paris (prenom.nom@cliffordchance.com)",
        "LinkedIn_Contact": "https://www.linkedin.com/company/clifford-chance-llp",
        "Site_Web": "https://www.cliffordchance.com",
        "Ticket_Moyen_Estime": "10 000 € - 25 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Convention de mécénat pour l'accès aux carrières juridiques",
        "Angle_Pitch_Ambition_Campus": "Préparation aux entretiens sélectifs et immersion dans un cabinet d'avocats de premier plan.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Contacter le comité Responsible Business au bureau de Paris (1 rue d'Astorg)."
    },
    {
        "ID": "ENT-18",
        "Nom_Entreprise": "Linklaters Paris",
        "Secteur_Activite": "Cabinet d'Avocats International",
        "Priorite": "Tier 2 - Droit & Juridique",
        "Nom_Contact": "Fondation Linklaters / Pôle RSE",
        "Poste_Contact": "Responsable Pédagogie Solidaire & Insertion",
        "Email_Contact": "via linklaters.com (25 rue de Marignan Paris)",
        "LinkedIn_Contact": "https://www.linkedin.com/company/linklaters",
        "Site_Web": "https://www.linklaters.com",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Partenariat Pédagogie Solidaire & Égalité des Chances",
        "Angle_Pitch_Ambition_Campus": "Insertion professionnelle des jeunes par l'apprentissage des codes et de l'éloquence.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Prendre contact avec l'équipe RSE au 25 rue de Marignan (Paris 8e)."
    },
    {
        "ID": "ENT-19",
        "Nom_Entreprise": "Google France",
        "Secteur_Activite": "Technologie & Numérique",
        "Priorite": "Tier 2 - Tech & Innovation",
        "Nom_Contact": "Direction des Relations Institutionnelles & RSE",
        "Poste_Contact": "Responsable Impact Social & Éducation",
        "Email_Contact": "contact via 8 rue de Londres 75009 Paris",
        "LinkedIn_Contact": "https://www.linkedin.com/company/google",
        "Site_Web": "https://about.google/intl/fr_fr/",
        "Ticket_Moyen_Estime": "20 000 € - 50 000 €",
        "Levier_Fiscal_60pct": "Don d'entreprise / Soutien philanthropique",
        "Type_Approche": "Mécénat financier + Accueil des visites lycéens dans leurs locaux",
        "Angle_Pitch_Ambition_Campus": "Nos lycéens visitent déjà le siège de Google France. Renforçons ce lien par un mécénat pérenne.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Proposer d'officialiser la convention de partenariat annuel (8 rue de Londres, Paris 9e)."
    },
    {
        "ID": "ENT-20",
        "Nom_Entreprise": "Microsoft France",
        "Secteur_Activite": "Technologie & Cloud",
        "Priorite": "Tier 2 - Tech & Innovation",
        "Nom_Contact": "Microsoft Philanthropies France",
        "Poste_Contact": "Directeur Philanthropies & Égalité des chances",
        "Email_Contact": "via 39 Quai du Président Roosevelt 92130 Issy",
        "LinkedIn_Contact": "https://www.linkedin.com/company/microsoft",
        "Site_Web": "https://www.microsoft.com/fr-fr",
        "Ticket_Moyen_Estime": "15 000 € - 35 000 €",
        "Levier_Fiscal_60pct": "Don 15 000 € = Coût net 6 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Programme Éducation & Mécénat d'équipement/financement",
        "Angle_Pitch_Ambition_Campus": "Équipement numérique et accompagnement des lycéens vers les carrières d'ingénieurs et de la tech.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Soumettre une demande sur le portail Microsoft pour les associations."
    },
    {
        "ID": "ENT-21",
        "Nom_Entreprise": "Amazon France",
        "Secteur_Activite": "E-Commerce, Cloud & Logistique",
        "Priorite": "Tier 3 - Tech & Logistique",
        "Nom_Contact": "Direction RSE & Programme Amazon Future Engineer",
        "Poste_Contact": "Responsable Programmes Éducation & Diversité",
        "Email_Contact": "via 67 bd du Général Leclerc 92110 Clichy",
        "LinkedIn_Contact": "https://www.linkedin.com/company/amazon",
        "Site_Web": "https://www.aboutamazon.fr",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Levier_Fiscal_60pct": "Don 15 000 € = Coût net 6 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Synergie avec Amazon Future Engineer & Bourses",
        "Angle_Pitch_Ambition_Campus": "Sensibilisation et propulsion des jeunes de banlieue vers les filières supérieures d'excellence.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Contacter l'équipe RSE à Clichy (92)."
    },
    {
        "ID": "ENT-22",
        "Nom_Entreprise": "Groupe TF1 (RSE & Marque Employeur)",
        "Secteur_Activite": "Médias & Télévision",
        "Priorite": "Tier 2 - Médias & Prise de Parole",
        "Nom_Contact": "Direction RSE & Diversité",
        "Poste_Contact": "Directrice RSE & Engagement des collaborateurs",
        "Email_Contact": "rse@tf1.fr / fondationtf1@tf1.fr",
        "LinkedIn_Contact": "https://www.linkedin.com/company/tf1",
        "Site_Web": "https://groupe-tf1.fr",
        "Ticket_Moyen_Estime": "8 000 € - 15 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Parrainage Concours d'Éloquence & Visites de plateaux",
        "Angle_Pitch_Ambition_Campus": "Modules Ethos/Pathos/Logos, concours annuel d'art oratoire et documentaire « Mérite sous condition ».",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Proposer aux journalistes/présentateurs de TF1 d'animer une masterclass éloquence."
    },
    {
        "ID": "ENT-23",
        "Nom_Entreprise": "L'Oréal France (Direction RSE)",
        "Secteur_Activite": "Cosmétique & Luxe",
        "Priorite": "Tier 2 - Grand Groupe",
        "Nom_Contact": "Direction RSE & Diversité France",
        "Poste_Contact": "Responsable Diversité & Égalité des Chances",
        "Email_Contact": "rse.france@loreal.com (01 47 56 70 00)",
        "LinkedIn_Contact": "https://www.linkedin.com/company/loreal",
        "Site_Web": "https://www.loreal.com/fr/",
        "Ticket_Moyen_Estime": "15 000 € - 30 000 €",
        "Levier_Fiscal_60pct": "Don 15 000 € = Coût net 6 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat d'émancipation & Parrainage de promotions",
        "Angle_Pitch_Ambition_Campus": "Propulser les talents féminins et masculins de banlieue vers les postes de cadres et dirigeants.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Prendre contact au siège de Clichy (41 rue Martre)."
    },
    {
        "ID": "ENT-24",
        "Nom_Entreprise": "LVMH (Moët Hennessy Louis Vuitton)",
        "Secteur_Activite": "Luxe & Métiers d'Excellence",
        "Priorite": "Tier 2 - Grand Groupe",
        "Nom_Contact": "Direction du Développement Durable & Mécénat",
        "Poste_Contact": "Responsable Mécénat & Diversité",
        "Email_Contact": "contact via 22 avenue Montaigne 75008 Paris",
        "LinkedIn_Contact": "https://www.linkedin.com/company/lvmh",
        "Site_Web": "https://www.lvmh.fr",
        "Ticket_Moyen_Estime": "20 000 € - 40 000 €",
        "Levier_Fiscal_60pct": "Don 20 000 € = Coût net 8 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat d'excellence académique & Bourses",
        "Angle_Pitch_Ambition_Campus": "Transmission des codes de l'excellence aux jeunes de banlieue accédant à Sciences Po et aux grandes écoles.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Adresser une proposition formelle au 22 avenue Montaigne."
    },
    {
        "ID": "ENT-25",
        "Nom_Entreprise": "Bouygues Construction / SA Bouygues",
        "Secteur_Activite": "BTP, Immobilier & Services",
        "Priorite": "Tier 3 - Grand Groupe Industriel",
        "Nom_Contact": "Direction RSE & Mécénat Territorial",
        "Poste_Contact": "Directeur RSE & Engagement Collaborateurs",
        "Email_Contact": "rse@bouygues-construction.com (01 30 60 33 00)",
        "LinkedIn_Contact": "https://www.linkedin.com/company/bouygues-construction",
        "Site_Web": "https://www.bouygues-construction.com",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat territorial & Parrainage de promotions",
        "Angle_Pitch_Ambition_Campus": "Fort ancrage dans les quartiers en rénovation urbaine où se situent les 36 lycées partenaires.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Contacter le siège Challenger à Guyancourt (78)."
    },
    {
        "ID": "ENT-26",
        "Nom_Entreprise": "Saint-Gobain France",
        "Secteur_Activite": "Matériaux & Industrie",
        "Priorite": "Tier 3 - Grand Groupe Industriel",
        "Nom_Contact": "Direction RSE & Mécénat France",
        "Poste_Contact": "Responsable Engagements Solidaires",
        "Email_Contact": "rse.france@saint-gobain.com",
        "LinkedIn_Contact": "https://www.linkedin.com/company/saint-gobain",
        "Site_Web": "https://www.saint-gobain.com/fr",
        "Ticket_Moyen_Estime": "10 000 € - 20 000 €",
        "Levier_Fiscal_60pct": "Don 10 000 € = Coût net 4 000 € (Art. 238 bis CGI)",
        "Type_Approche": "Mécénat financier + Parrainage de collaborateurs",
        "Angle_Pitch_Ambition_Campus": "Ascension sociale par les études supérieures : un modèle 100% bénévole gage d'efficacité maximale.",
        "Statut_Prospection": "À contacter",
        "Notes_Action": "Cibler la Tour Saint-Gobain à Paris La Défense."
    }
]


def clean_text_field(text):
    """Nettoie les artefacts de recherche web (+1, +2, sites collés, etc.)."""
    if not text or pd.isna(text):
        return ""
    t = str(text)
    # Nettoyer les balises de citations web (+1, +2, etc.)
    t = re.sub(r'(\w+[\.\-\w]*)\+\d+', r'\1', t)
    # Nettoyer les domaines cités collés (wikipedia, sia-partners, etc.)
    t = re.sub(r'\b(wikipedia|carrieresfrance\.[a-z]+|adlittle|oliverwyman|kearney|simon-kucher|capgemini|annuaire-entreprises\.data\.gouv|droitsdurgence|darrois|officiel-inclusion|eurazeo|tikehaucapital|ardian|amundi|oddo-bhf|climate-transparency-hub\.ademe|allianz|groupama|food\.ec\.europa|se)\b', '', t, flags=re.IGNORECASE)
    # Nettoyer les suffixes d'URL collés
    t = re.sub(r'(https?://[^\s\)]+)(sia-partners|carrieresfrance|adlittle|oliverwyman|kearney|simon-kucher|capgemini|eurazeo|tikehaucapital|ardian|amundi|oddo-bhf|allianz|groupama|danone|schneider)', r'\1', t)
    # Espaces multiples
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def load_and_merge_entreprises():
    """Charge le fichier téléchargé, le nettoie et le fusionne avec le lot initial."""
    all_data = list(BASE_ENTREPRISES_LOT1)

    if os.path.exists(DOWNLOAD_PATH):
        print(f"Chargement des nouvelles entreprises depuis : {DOWNLOAD_PATH}")
        with open(DOWNLOAD_PATH, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.DictReader(f)
            for row in reader:
                eid = clean_text_field(row.get("ID", ""))
                # Ne pas dupliquer si déjà présent
                if any(x["ID"] == eid for x in all_data):
                    continue

                nom = clean_text_field(row.get("Nom_Entreprise", ""))
                secteur = clean_text_field(row.get("Secteur_Activite", ""))
                contact = clean_text_field(row.get("Nom_Contact_Decideur", ""))
                poste = clean_text_field(row.get("Poste_Contact", ""))
                email = clean_text_field(row.get("Email_Contact_Verifie", ""))
                source = clean_text_field(row.get("Source_Officielle_RSE (Lien/Rapport)", ""))
                ticket = clean_text_field(row.get("Ticket_Moyen_Estime", ""))
                cout_net = clean_text_field(row.get("Cout_Net_Apres_IS_60pct", ""))
                rse_prog = clean_text_field(row.get("Programme_RSE_Existant", ""))
                pitch = clean_text_field(row.get("Angle_Pitch_Ambition_Campus", ""))

                # Extraction d'une URL propre depuis la source RSE
                url_match = re.search(r'https?://[^\s\)]+', source)
                site_web = url_match.group(0) if url_match else f"https://www.{nom.lower().replace(' ', '').replace('’', '').replace('france', '')}.com"

                # Attribution de la priorité selon le profil
                if any(k in nom.lower() for k in ["sia partners", "bearingpoint", "oliver wyman", "roland berger", "bredin prat", "darrois", "white & case"]):
                    prio = "Tier 2 - Cabinet de Référence (Conseil / Droit)"
                elif any(k in nom.lower() for k in ["eurazeo", "tikehau", "ardian", "amundi", "oddo", "blackrock", "allianz", "groupama"]):
                    prio = "Tier 2 - Finance, Private Equity & Assurance"
                else:
                    prio = "Tier 3 - Grand Groupe & Entreprise à Mission"

                item = {
                    "ID": eid,
                    "Nom_Entreprise": nom,
                    "Secteur_Activite": secteur,
                    "Priorite": prio,
                    "Nom_Contact": contact,
                    "Poste_Contact": poste,
                    "Email_Contact": email if email and "À compléter" not in email else f"via direction RSE ({site_web})",
                    "LinkedIn_Contact": f"https://www.linkedin.com/company/{nom.lower().replace(' ', '-').replace('’', '')}",
                    "Site_Web": site_web,
                    "Ticket_Moyen_Estime": ticket,
                    "Levier_Fiscal_60pct": f"Don {ticket} = Coût net {cout_net} (Déduction 60% IS Art. 238 bis CGI)",
                    "Type_Approche": f"Convention Mécénat / Partenariat {rse_prog[:40]}",
                    "Angle_Pitch_Ambition_Campus": pitch,
                    "Statut_Prospection": "À qualifier & contacter",
                    "Notes_Action": f"Proposer une convention de mécénat 2026-2027 et masterclass métier. Source RSE : {source[:60]}"
                }
                all_data.append(item)

    return all_data


def generate_database():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    entreprises = load_and_merge_entreprises()
    df = pd.DataFrame(entreprises)

    # 1. Export CSV propre encodé UTF-8 BOM
    df.to_csv(CSV_PATH, index=False, encoding="utf-8-sig", sep=";")
    print(f"[OK] Master CSV Entreprises genere : {CSV_PATH} ({len(df)} entreprises au total)")

    # 2. Export Excel formaté professionnellement
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Master_Entreprises", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Master_Entreprises"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F1E36", end_color="0F1E36", fill_type="solid")
        cell_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        col_widths = {
            "ID": 10,
            "Nom_Entreprise": 28,
            "Secteur_Activite": 28,
            "Priorite": 26,
            "Nom_Contact": 26,
            "Poste_Contact": 30,
            "Email_Contact": 32,
            "LinkedIn_Contact": 30,
            "Site_Web": 30,
            "Ticket_Moyen_Estime": 20,
            "Levier_Fiscal_60pct": 35,
            "Type_Approche": 32,
            "Angle_Pitch_Ambition_Campus": 38,
            "Statut_Prospection": 24,
            "Notes_Action": 35,
        }

        for row_idx in range(2, len(df) + 2):
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if col_name == "Priorite":
                    val = str(cell.value)
                    if "Tier 1" in val:
                        cell.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="166534")
                    elif "Tier 2" in val:
                        cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="1E40AF")
                    elif "Tier 3" in val:
                        cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="92400E")

        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = col_widths.get(col_name, 22)

        worksheet.freeze_panes = "A2"

    print(f"[OK] Master Excel Entreprises genere : {XLSX_PATH}")


if __name__ == "__main__":
    generate_database()
